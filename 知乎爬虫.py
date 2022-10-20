from apscheduler.schedulers.blocking import BlockingScheduler
import requests
import json
import datetime
import openpyxl
import os
sched = BlockingScheduler()
#定时任务
@sched.scheduled_job('cron', hour='0-23')
def timed_job():
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'}
    def getTLSt(n):
        title=[]
        link=[]
        score=[]
        topics=[]
        answer_num=[]
        new_answer_num=[]
        follow_num=[]
        new_follow_num=[]
        pv=[]
        new_pv=[]
        upvote_num=[]
        new_upvote_num=[]
        for m in range(2):
            url='https://www.zhihu.com/api/v4/creators/rank/hot?domain=1000'+str(n).rjust(2,'0')+'&limit=20&offset='+str(m*20)+'&period=hour'
            json=requests.get(url=url,headers=headers).json()
            for i in range(len(json['data'])):
                title.append(json['data'][i]['question']['title'])
                link.append(json['data'][i]['question']['url'])
                score.append(json['data'][i]['reaction']['score'])
                answer_num.append(json['data'][i]['reaction']['answer_num'])
                new_answer_num.append(json['data'][i]['reaction']['new_answer_num'])
                follow_num.append(json['data'][i]['reaction']['follow_num'])
                new_follow_num.append(json['data'][i]['reaction']['new_follow_num'])
                pv.append(json['data'][i]['reaction']['pv'])
                new_pv.append(json['data'][i]['reaction']['new_pv'])
                upvote_num.append(json['data'][i]['reaction']['upvote_num'])
                new_upvote_num.append(json['data'][i]['reaction']['new_upvote_num'])
                topic_single=[]
                for j in range(len(json['data'][i]['question']['topics'])):
                    topic_single.append(json['data'][i]['question']['topics'][j]['name'])
                topics.append(topic_single)
        return title,link,score,topics,answer_num,new_answer_num,follow_num,new_follow_num,pv,new_pv,upvote_num,new_upvote_num
    url='https://www.zhihu.com/api/v4/creators/domain'
    json=requests.get(url=url,headers=headers).json()
    cla=['全部']
    for i in range(len(json['domains'])):
        for j in range(len(json['domains'][i]['items'])):
            cla.append(json['domains'][i]['items'][j]['name'])
    ti = datetime.datetime.now()
    time_hot = str(ti.month) + '月' + str(ti.day) + '日 ' + str(ti.hour) + '时' + str(ti.minute) + '分'
    time_file = str(ti.month) + '月' + str(ti.day) + '日 ' + str(ti.hour) + '时'
    file_path = './知乎热榜_Second' + time_hot + '.xlsx'
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)
    wb = openpyxl.load_workbook(file_path)
    for j in range(len(cla)):
        ws = wb.create_sheet(title=cla[j])
        title,link,score,topics,answer_num,new_answer_num,follow_num,new_follow_num,pv,new_pv,upvote_num,new_upvote_num=getTLSt(j)
        ws.cell(row=1, column=1, value='排名')
        ws.cell(row=1, column=2, value='链接')
        ws.cell(row=1, column=3, value='问题')
        ws.cell(row=1, column=4, value='热力值')
        ws.cell(row=1, column=5, value='热点分类')
        ws.cell(row=1, column=6, value='回答数')
        ws.cell(row=1, column=7, value='新回答数')
        ws.cell(row=1, column=8, value='关注数')
        ws.cell(row=1, column=9, value='新关注数')
        ws.cell(row=1, column=10, value='浏览数')
        ws.cell(row=1, column=11, value='新浏览数')
        ws.cell(row=1, column=12, value='赞同数')
        ws.cell(row=1, column=13, value='新赞同数')
        for i in range(len(title)):
            ws.cell(row=i+2, column=1, value=i+1)
            ws.cell(row=i+2, column=3, value=title[i])
            ws.cell(row=i+2, column=2, value=link[i])
            ws.cell(row=i+2, column=4, value=score[i])
            ws.cell(row=i+2, column=5, value=','.join(topics[i]))
            ws.cell(row=i+2, column=6, value=answer_num[i])
            ws.cell(row=i+2, column=7, value=new_answer_num[i])
            ws.cell(row=i+2, column=8, value=follow_num[i])
            ws.cell(row=i+2, column=9, value=new_follow_num[i])
            ws.cell(row=i+2, column=10, value=pv[i])
            ws.cell(row=i+2, column=11, value=new_pv[i])
            ws.cell(row=i+2, column=12, value=upvote_num[i])
            ws.cell(row=i+2, column=13, value=new_upvote_num[i])
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)
    wb.close()
sched.start()